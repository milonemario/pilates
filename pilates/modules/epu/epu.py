"""
Module to provide processing functions for EPU data.

EPU data refers to the Economic Policy Uncertainty data from
Baker, Bloom and Davis (2016).
https://www.policyuncertainty.com

"""
import os
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
from pilates import data_module
from openpyxl import load_workbook


class epu(data_module):

    def __init__(self, d):
        data_module.__init__(self, d)

    def _convert_data(self, name, path, force=False,
                     delim_whitespace=False, nrows=None):
        if name not in ['country']:
            data_module._convert_data(self, name, path)
        else:
            # For epu country, the Excel file is not clean.
            # Need to find how many rows to consider.
            wb = load_workbook(path)
            ws = wb.active
            df = pd.DataFrame(ws.values)
            # Find last row with year info
            row = 430
            while (str(type(ws['A'+str(row)].value))=="<class 'int'>"):
                row += 1
            row -= 1
            data_module._convert_data(self, name, path, nrows=row)

    def add_file_old(self, name, path, nrows=None, force=False, types=None):
        """ Add a file to the module.
        Converts and make the file available for the module to use.

        Args:
            name (str): Name of the file to be used by the module.
            path (str): Path of the original file to convert.
            force (bool, optional): If False, the file is not re-converted.
                Defaults to False.
            types (str, optional): Path of the Yaml file containing the fields
                types of the data.

        """

        self.d._check_data_dir()
        # Get the file type
        filename, ext = os.path.splitext(os.path.basename(path))
        # Create the new file name
        filename_pq = self.d.datadir+filename+'.parquet'
        # Check if the file has already been converted
        if os.path.exists(filename_pq) and force is False:
            None
        else:
            # Open the file (EPU uses excel files)
            df = pd.read_excel(path, nrows=nrows)

            # Write the data
            df.columns = map(str.lower, df.columns)  # Lower case col names
            t = pa.Table.from_pandas(df)
            pqwriter = pq.ParquetWriter(filename_pq, t.schema)
            pqwriter.write_table(t)
            pqwriter.close()
        # Link the name to the actual filename
        setattr(self, name, filename)

    def set_level(self, level):
        """ Define the aggregation level of the index.
        It can be 'country' or 'firm'
        """
        if level in ['country']:
            self.add_file('country')
            self.epu = self.country
            self.level = level
        elif level in ['firm']:
            self.add_file('firm')
            self.epu = self.firm
            self.level = level
        else:
            raise Exception('EPU aggregation level should by either',
                            'country or firm')

    def set_maindata(self, maindata):
        """ Set the main data used in the study (for merging purposes).
        Could be 'comp' for compustat of 'crps' for CRSP.
        """
        if maindata in ['crsp', 'comp']:
            self.maindata = maindata
            self.maindata = maindata
        else:
            raise Exception('EPU main data should by either',
                            'comp or crsp')

    def set_keys(self):
        if self.level == 'country':
            self.key = ['year', 'month']
            self.data_key = [self.d.col_date]
        elif self.level == 'firm':
            if self.maindata == 'comp':
                self.key = [self.d.comp.col_id, 'year']
                self.data_key = [self.d.comp.col_id, self.d.col_date]
            elif self.maindata == 'crsp':
                self.key = [self.d.crsp.col_id, 'year']
                self.data_key = [self.d.crsp.col_id, self.d.col_date]

    def get_fields(self, data, fields):
        self.set_keys()
        df = self.open_data(self.epu, self.key+fields)
        if self.level == 'firm':
            # 2 duplicates on that file at locs [29196, 52509]
            df = df.drop([29196, 52509])
        # Merge to the user data
        dfu = self.open_data(data, self.data_key)
        # Extract year and month from user data
        dt = pd.to_datetime(dfu[self.d.col_date]).dt
        dfu['year'] = dt.year
        dfu['month'] = dt.month
        # Merge
        dfin = dfu.merge(df, how='left', on=self.key)
        dfin.index = dfu.index
        return(dfin[fields])
